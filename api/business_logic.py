from typing import Any, Dict, List, Literal, Tuple
from db_models import (
    Role,
    User,
    UserDetail,
    PaymentAccount,
    Contractor,
    ContractType,
    RepeatingContract,
    Contract,
    Invoice,
    DebitCredit,
    UserActionsLog,
    IT_IS_TYPES,
)
import numpy as np
from my_engine import Session
from sqlalchemy import func, select, text, or_, and_
from sqlalchemy import Column, Table
from sqlalchemy.orm.query import Query
from datetime import date, datetime
from cachetools import TTLCache
import secrets
import my_exceptions as my_exc
from openpyxl import Workbook
from openpyxl.styles import numbers
from io import BytesIO

query_cache = TTLCache(100, 14400)
LAST_DAY_IN_MONTHS = dict()



def append_filters(stmt: Query, args: dict):
    """Добавление фильтров для запроса списка контрактов"""
    # Contract start_date stop_date
    is_invoice_not_joined = 'invoices' not in [_[0].name for _ in stmt._legacy_setup_joins]


    if (value := args.get("contract_id")) and args.get("contract_id") != 0:
        return stmt.filter(Contract.id == value)


    if value := args.get("start_report_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(or_(
            Contract.repeating_month >= value,
            Invoice.payment_date >= value,
            Invoice.payment_expected_date >= value
        ))
    if value := args.get("end_report_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(or_(
            Contract.repeating_month <= value,
            Invoice.payment_date <= value,
            Invoice.payment_expected_date <= value
        ))


    if value := args.get("it_is"):
        stmt = stmt.filter(Contract.it_is == value)
    if value := args.get("contractor_id"):
        stmt = stmt.filter(Contract.contractor_id == value)
    if value := args.get("description"):
        stmt = stmt.filter(Contract.description.like(f"%{value}%"))
    if value := args.get("type_id"):
        stmt = stmt.filter(Contract.type_id == value)
    if value := args.get("start_contract_date"):
        stmt = stmt.filter(Contract.contract_date >= value)
    if value := args.get("end_contract_date"):
        stmt = stmt.filter(Contract.contract_date <= value)
    if value := args.get("contract"):
        stmt = stmt.filter(Contract.contract.like(f"%{value}%"))
    if value := args.get("order"):
        stmt = stmt.filter(Contract.order.like(f"%{value}%"))
    if value := args.get("start_order_date"):
        stmt = stmt.filter(Contract.order_date >= value)
    if value := args.get("end_order_date"):
        stmt = stmt.filter(Contract.order_date <= value)
    if value := args.get("start_order_deadline"):
        stmt = stmt.filter(Contract.order_deadline >= value)
    if value := args.get("end_order_deadline"):
        stmt = stmt.filter(Contract.order_deadline <= value)
    if value := args.get("order_price"):
        stmt = stmt.filter(Contract.order_price == value)
    if value := args.get("start_repeating_contract"):
        stmt = stmt.filter(Contract.repeating_month >= value)
    if value := args.get("end_repeating_contract"):
        stmt = stmt.filter(Contract.repeating_month < value)


    if value := args.get("invoice_name"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.name.like(f"%{value}%"))
    if value := args.get("start_invoice_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.date >= value)
    if value := args.get("end_invoice_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.date <= value)
    if value := args.get("invoice_price"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.price == value)
    if value := args.get("payment_amount"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.payment_amount == value)
    if value := args.get("start_payment_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.payment_date >= value)
    if value := args.get("end_payment_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.payment_date <= value)
    if value := args.get("start_payment_expected_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.payment_expected_date >= value)
    if value := args.get("end_payment_expected_date"):
        if is_invoice_not_joined:
            is_invoice_not_joined = False
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.filter(Invoice.payment_expected_date <= value)
    return stmt


def append_order(stmt: Query, name: str, order_type: str):
    joined_tables = [_[0].name for _ in stmt._legacy_setup_joins]
    column: Column
    def asc_desc(column: Column):
        if order_type.upper() == 'DESC':
            return column.desc()
        if order_type.upper() == 'ASC':
            return column.asc()
        return None
    if name in (
        '', 
        'invoice_name', 
        'invoice_date', 
        'invoice_price', 
        'payment_amount', 
        'payment_date', 
        'payment_expected_date',
    ):
        stmt = stmt.order_by(Contract.id.desc())
    elif name == 'it_is':
        stmt = stmt.order_by(asc_desc(Contract.it_is))
    elif name == 'contractor_id':
        if 'contractor' not in joined_tables:
            stmt = stmt.join(Contractor)
        stmt = stmt.order_by(asc_desc(Contractor.name))
    elif name == 'description':
        stmt = stmt.order_by(asc_desc(Contract.description))
    elif name == 'type_id':
        if 'contract_types' not in joined_tables:
            stmt = stmt.join(ContractType)
        stmt = stmt.order_by(asc_desc(ContractType.name))
    elif name == 'contract_date':
        stmt = stmt.order_by(asc_desc(Contract.contract_date))
    elif name == 'contract':
        stmt = stmt.order_by(asc_desc(Contract.contract))
    elif name == 'order':
        stmt = stmt.order_by(asc_desc(Contract.order))
    elif name == 'order_date':
        stmt = stmt.order_by(asc_desc(Contract.order_date))
    elif name == 'order_price':
        stmt = stmt.order_by(asc_desc(Contract.order_price))
    elif name == 'invoice_name':
        if 'invoices' not in joined_tables:
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.order_by(asc_desc(Invoice.name))
    elif name == 'invoice_date':
        if 'invoices' not in joined_tables:
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.order_by(asc_desc(Invoice.date))
    elif name == 'invoice_price':
        if 'invoices' not in joined_tables:
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.order_by(asc_desc(Invoice.price))
    elif name == 'payment_amount':
        if 'invoices' not in joined_tables:
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.order_by(asc_desc(Invoice.payment_amount))
    elif name == 'payment_date':
        if 'invoices' not in joined_tables:
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.order_by(asc_desc(Invoice.payment_date))
    elif name == 'payment_expected_date':
        if 'invoices' not in joined_tables:
            stmt = stmt.join(Invoice, isouter=True)
        stmt = stmt.order_by(asc_desc(Invoice.payment_expected_date))
    return stmt


def get_contracts(filters: dict):
    limit_rows = 100
    query_token = filters.get('query_token')
    sess = Session()
    try:
        contracts: Query = None
        if query_token:
            contracts = query_cache.get(query_token)
        if contracts is None:
            contracts = sess.query(Contract).filter(Contract.is_active == 1).order_by(Contract.id.desc())
            if filters:
                contracts = append_filters(contracts, filters)
            query_token = secrets.token_hex(16)
        if 'sort_key' in filters:
            contracts = append_order(contracts.order_by(None), filters['sort_key'], filters['sort_order'])
        query_cache[query_token] = contracts
        row_count = contracts.count()
        page_count_stmt = row_count // limit_rows + (1 if row_count % limit_rows else 0)
        contracts = contracts.offset(limit_rows * (filters.get('page', 1) - 1))
        contracts = contracts.limit(limit_rows)

        result = [_.to_dict() for _ in contracts]
    except Exception as e:
        raise e
    finally:
        sess.close()
    return result, 200, ('query_token', query_token), ('page_counts', page_count_stmt), ('row_counts', row_count), ('limit_rows', limit_rows)


def get_contact_details(contract_id, is_main_contracts = True):
    sess = Session()
    try:
        contract: Contract | None = sess.query(Contract).filter(Contract.id == contract_id).one_or_none()
        if contract is None:
            return None, 404
        result = contract.to_dict()
        result['repeating_contract'] = contract.fk__repeating_contract.to_dict() if contract.repeating_contract_id else {"start_date": result['repeating_month']}
        if is_main_contracts:
            if contract.fk__debit:
                result['linked_contracts'] = [get_contact_details(_.credit, False) for _ in contract.fk__debit]
            else:
                result['linked_contracts'] = [get_contact_details(_.debit, False) for _ in contract.fk__credit]
    except Exception as e:
        raise e
    finally:
        sess.close()
    
    if is_main_contracts:
        return result, 200
    return result


def get_lists():
    sess = Session()
    result: Dict[str, List[Dict[str, Any] | None]] = dict()
    try:
        contractors = sess.query(Contractor)
        result['contractors'] = [_.to_dict() for _ in contractors]
        
        contract_types = sess.query(ContractType)
        result['contract_type'] = [_.to_dict() for _ in contract_types]
    except Exception as e:
        raise e
    finally:
        sess.close()
    return result, 200


def __check_last_day_in_month(repeat_month: str, day: str):
    """Проверяет возвращает строку с датой"""
    year_month, _ = repeat_month.rsplit('-', 1)

    if not (last_day_in_month := LAST_DAY_IN_MONTHS.get(year_month)):
        last_day_in_month = (np.datetime64(year_month) + np.timedelta64(1, 'M') - np.timedelta64(1, 'D')).astype(date).day
        LAST_DAY_IN_MONTHS[year_month] = last_day_in_month

    return f"{year_month}-{day}" if int(day) <= last_day_in_month else f"{year_month}-{last_day_in_month}"


def __insert_log(user, section, text):
    with Session() as sess:
        sess.add(UserActionsLog(user_id=user, section=section, changes=text))
        sess.commit()



def insert_contract(contract_args: dict, user):


    repeating_contract_args: dict = contract_args.pop('repeating_contract', {})
    invoices_args: List[dict] = contract_args.pop('invoice', [])
    linked_contracts_args: List[dict] = contract_args.pop('linked_contracts', [])
    contract_args['user_id'] = user
    


    repeating_contract_args.pop('id', None)
    contract_args.pop('id', None)
    for i in invoices_args:
        i.pop('id', None)

    if Contract.assert_args(contract_args):
        contract = Contract(**contract_args)
    else:
        return {'message': 'Некорректные данные в блоке "Договор" или "Заказ"'}, 400

    sess = Session()
    try:
        current_date = date.today().replace(day=1)
        repeating_contract = None

        if repeating_contract_args.get('end_date') and RepeatingContract.assert_args(repeating_contract_args):
            if repeating_contract_args.get('start_date') and repeating_contract_args.get('end_date'):
                if repeating_contract_args.get('start_date') >= repeating_contract_args.get('end_date'):
                    return {'message': 'Ошибка в датах. Дата начала должна быть меньше даты конца'}, 400

                start_date = np.datetime64(repeating_contract_args['start_date'])
                end_date = np.datetime64(repeating_contract_args['end_date'])
                if start_date < current_date:
                    return {'message': 'Ошибка в датах. Дата начала должна быть больше или равна текущей даты'}, 400
                repeating_contract_args['start_date'] = start_date.astype(date)
                repeating_contract_args['end_date'] = end_date.astype(date)
                repeating_dates = np.arange(start_date + np.timedelta64('1', 'M'), end_date + np.timedelta64('1', 'M'), np.timedelta64('1', 'M')).astype(date)
            else:
                raise my_exc.DateNotEnough

            if repeating_contract_args.get('start_pause_date') and repeating_contract_args.get('end_pause_date'):
                start_pause_date = np.datetime64(repeating_contract_args['start_pause_date'])
                end_pause_date = np.datetime64(repeating_contract_args['end_pause_date'])
                repeating_contract_args['start_pause_date'] = start_pause_date.astype(date)
                repeating_contract_args['end_pause_date'] = end_pause_date.astype(date)

                if start_pause_date >= end_pause_date:
                    return {'message': 'Ошибка в датах. Дата начала должна быть меньше даты конца'}, 400
                if start_pause_date <= start_date:
                    return {'message': 'Ошибка в датах. Дата начала исключения меньше или равна дате начала повторения'}, 400
                if end_pause_date >= end_date:
                    return {'message': 'Ошибка в датах. Дата конца исключения больше или равна дате конца повторения'}, 400

                excluded_repeating_dates = np.arange(start_pause_date, end_pause_date + np.timedelta64('1', 'M'), np.timedelta64('1', 'M')).astype(datetime)
            elif repeating_contract_args.get('start_pause_date') or repeating_contract_args.get('end_pause_date'):
                raise my_exc.DateNotEnough
            else:
                excluded_repeating_dates = list()

            repeating_contract = RepeatingContract(**repeating_contract_args)
            sess.add(repeating_contract)
            sess.commit()
            contract.repeating_contract_id = repeating_contract.id
        elif repeating_contract_args.get('start_date') and RepeatingContract.assert_args(repeating_contract_args):
            start_date = np.datetime64(repeating_contract_args['start_date'])
            if start_date < current_date:
                return {'message': 'Ошибка в датах. Месяц платежа должен быть больше или равен текущей даты'}, 400
            contract.repeating_month = start_date.astype(date)
        else:
            raise my_exc.DateNotEnough
    except my_exc.DateNotEnough:
        sess.close()
        if not RepeatingContract.assert_args(repeating_contract_args):
            return {'message': 'Месяц платежа. год необходимо указывать полностью'}, 400
        return {'message': 'Месяц платежа. Ошибка в дате'}, 400
    except ValueError:
        sess.close()
        return {'message': 'Месяц платежа. Ошибка в дате'}, 400
    except Exception as e:
        sess.close()
        raise e
    
    try:
        contracts_list = [contract]
        if repeating_contract:
            contract.repeating_month = start_date.astype(date)
            for d in repeating_dates:
                if d in excluded_repeating_dates:
                    continue
                contracts_list.append(
                    Contract(
                        repeating_month = d,
                        repeating_contract_id = contract.repeating_contract_id,
                        **contract_args
                    )
                )

        sess.add_all(contracts_list)
        sess.commit()
        log_text = f'Создан контракт:<br>  ID: {str([_.id for _ in contracts_list])[1:-1]}'
        contract_id = contract.id
    except Exception as e:
        sess.rollback()
        if repeating_contract:
            sess.delete(repeating_contract)
            sess.commit()
        sess.close()
        raise e
    
    try:
        if invoices_args:
            invoices = [Invoice(contract_id = contract_id, **_) for _ in invoices_args if Invoice.assert_args(_)]

            if len(invoices) != len(invoices_args):
                sess.rollback()
                for c in contracts_list:
                    sess.delete(c)
                if repeating_contract:
                    sess.delete(repeating_contract)
                sess.commit()
                sess.close()
                return {"message": "Ошибка указания дат в блоке счетов"}, 400

            
            if len(contracts_list) > 1:
                for c in contracts_list[1:]:
                    repeating_month: date = c.repeating_month
                    invoices.extend([
                        Invoice(
                            contract_id = c.id,
                            price = _['price'],
                            payment_expected_date = __check_last_day_in_month(
                                repeating_month.strftime("%Y-%m-%d"),
                                _['payment_expected_date'].rsplit('-', 1)[1]
                            )
                        ) 
                        for _ in invoices_args
                    ])

            sess.add_all(invoices)
            sess.commit()
    except Exception as e:
        sess.rollback()
        for c in contracts_list:
            sess.delete(c)
        if repeating_contract:
            sess.delete(repeating_contract)
        sess.commit()
        sess.close()
        raise e
    
    try:
        if linked_contracts_args:
            linked_contracts = (insert_contract(_, user) for _ in linked_contracts_args)
            if contract.it_is == {'Доход'}:
                debit_credit = [DebitCredit(debit=contract_id, credit=_[0]) for _ in linked_contracts if _[1] == 201]
            else:
                debit_credit = [DebitCredit(debit=_[0], credit=contract_id) for _ in linked_contracts if _[1] == 201]
            if len(debit_credit) != len(linked_contracts_args):
                raise 
            sess.add_all(debit_credit)
            sess.commit()
    except Exception as e:
        sess.rollback()
        for i in invoices:
            sess.delete(i)
        for c in contracts_list:
            sess.delete(c)
        if repeating_contract:
            sess.delete(repeating_contract)
        sess.commit()
        sess.close()
        raise e
    sess.close()

    __insert_log(user, 'Доход/Расход', log_text)

    return contract_id, 201


def update_contract(contract_id, contract_args: dict, user):

    repeating_contract_args: dict = contract_args.pop('repeating_contract', {}) or {}
    invoices_args: List[dict] = contract_args.pop('invoice', []) or []
    linked_contracts_args: List[dict] = contract_args.pop('linked_contracts', []) or []
    

    if not Contract.assert_args(contract_args):
        return {'message': 'Ошибка в блоке "Договор" или "Заказ"'}, 400
    
    if not RepeatingContract.assert_args(repeating_contract_args):
        return {'message': 'Месяц платежа. Год необходимо указывать полностью'}, 400

    current_date = date.today().replace(day=1)
    sess = Session()
    try:
        contract: Contract = sess.get(Contract, contract_id)
        if contract is None:
            return {"message": "Контракт не найден.<br>Пожалуйста обновите страницу"}, 404

        log_with_changes = f'Изменены данные у контракта с ID {contract.id} :'

        if contract.repeating_contract_id and not repeating_contract_args:
            #Удаление ежемесячного повторения
            log_with_changes += '<br>  - Снята галка "Ежемесячный"'
            repeating_contract: RepeatingContract = contract.fk__repeating_contract
            not_overdue_contracts: List[Contract] = sess.query(Contract).filter(
                Contract.repeating_contract_id == contract.repeating_contract_id,
                Contract.repeating_month >= current_date,
                Contract.id != contract.id,
            )
            for c in not_overdue_contracts:
                for i in c.fk__invoice:
                    sess.delete(i)
                if c.it_is == {'Доход'}:
                    dc_list = c.fk__debit
                else:
                    dc_list = c.fk__credit
                for dc in dc_list:
                    sess.delete(dc)
                sess.delete(c)
            log_with_changes

            overdue_contracts: List[Contract] = sess.query(Contract).filter(
                Contract.repeating_contract_id == contract.repeating_contract_id,
                Contract.repeating_month < current_date,
                Contract.id != contract.id,
            )
            for c in overdue_contracts:
                c.repeating_contract_id = None
            
            contract.repeating_contract_id = None
            sess.delete(repeating_contract)
            sess.commit()
        elif not contract.repeating_contract_id and repeating_contract_args.get('end_date'):
            contract_pattern = contract_args.copy()
            old_contract_id = contract_pattern.pop('id')
            contract_pattern['repeating_contract'] = repeating_contract_args

            new_contract_id, status = insert_contract(contract_pattern, user)
            log_with_changes += f'<br>  - поставлена галка "Ежемесячный". Изменён ID контракта с {old_contract_id} на {new_contract_id}'
            if status == 201:
                i: Invoice
                for i in contract.fk__invoice:
                    i.contract_id = new_contract_id
                    sess.commit()
                
                dc: DebitCredit
                if contract.it_is == {'Доход'}:
                    for dc in contract.fk__debit:
                        dc.debit = new_contract_id
                        sess.commit()
                else:
                    for dc in contract.fk__credit:
                        dc.credit = new_contract_id
                        sess.commit()
            else:
                return new_contract_id, status
            sess.delete(contract)
            sess.commit()
            contract_id = new_contract_id
            contract: Contract = sess.get(Contract, contract_id)
        elif repeating_contract_args.get('start_date') and not repeating_contract_args.get('end_date'):
            new_repeating_month: date = np.datetime64(repeating_contract_args['start_date']).astype(date)
            log_old_date = contract.repeating_month.strftime('%m.%Y') if contract.repeating_month else 'пусто'
            log_new_date = new_repeating_month.strftime('%m.%Y')
            if log_old_date != log_new_date:
                contract.repeating_month = new_repeating_month
                log_with_changes += f'<br>  - "Месяц платежа". Было {log_old_date} стало {log_new_date}'

        
        human_date = "%d.%m.%Y"

        contract.get_difference(contract_args)
        if value := contract_args.get("contractor_id"):
            old_name = contract.fk__contractor.name

            contract.contractor_id = value
            sess.commit()
            new_name = contract.fk__contractor.name

            log_with_changes += f'<br>  - "Контрагент". Было {old_name} стало {new_name}'
        
        if value := contract_args.get("type_id"):
            it_is = [*contract.it_is][0].lower()

            old_name = contract.fk__type.name
            contract.type_id = value
            sess.commit()
            new_name = contract.fk__type.name

            log_with_changes += f'<br>  - "Тип {it_is}а". Было {old_name} стало {new_name}'

        if value := contract_args.get("description"):
            log_with_changes += f'<br>  - "Описание". Было {contract.description} стало {value}'
            contract.description = value

        if value := contract_args.get("contract"):
            old_value = contract.contract or 'пустым'
            log_with_changes += f'<br>  - "Номер договора". Было {old_value} стало {value}'
            contract.contract = value
        
        if value := contract_args.get("contract_date"):
            new_value = datetime.strptime(value, '%Y-%m-%d').strftime(human_date)
            old_value = contract.contract_date.strftime(human_date) if contract.contract_date else 'пустой'
            log_with_changes += f'<br>  - "Дата договора". Было {old_value} стало {new_value}'
            contract.contract_date = value

        if value := contract_args.get("comment"):
            old_value = contract.comment or 'пустым'
            log_with_changes += f'<br>  - "Комментарий". Было {old_value} стало {value}'
            contract.comment = value

        if value := contract_args.get("order"):
            old_value = contract.order or 'пустым'
            log_with_changes += f'<br>  - "Номер заказа". Было {old_value} стало {value}'
            contract.order = value

        if value := contract_args.get("order_date"):
            new_value = datetime.strptime(value, '%Y-%m-%d').strftime(human_date)
            old_value = contract.order_date.strftime(human_date) if contract.order_date else 'пустой'
            log_with_changes += f'<br>  - "Дата заказа". Было {old_value} стало {new_value}'
            contract.order_date = value

        if value := contract_args.get("order_price"):
            log_with_changes += f'<br>  - "Общая стоимость, с НДС". Было {contract.order_price} стало {value}'
            contract.order_price = value

        if value := contract_args.get("order_deadline"):
            new_value = datetime.strptime(value, '%Y-%m-%d').strftime(human_date)
            old_value = contract.order_deadline.strftime(human_date)
            log_with_changes += f'<br>  - "Дата окончания заказа". Было {old_value} стало {new_value}'
            contract.order_deadline = value
        sess.commit()

    except ValueError:
        sess.close()
        return {'message': 'Ошибка в дате'}, 400
    except Exception as e:
        sess.close()
        raise e

    try:
        existing_invoices_in_args=list()
        for invoice_args in invoices_args:
            if not Invoice.assert_args(invoice_args):
                return {"message": "Ошибка в данных счёта."}, 400
            
            if invoice_args.get('id'):
                invoice: Invoice = sess.get(Invoice, invoice_args['id'])
                invoice.get_difference(invoice_args)

                if 'name' in invoice_args:
                    value = invoice_args['name']
                    old_value = invoice.name or 'пустым'
                    new_value = value or 'пустым'
                    log_with_changes += f'<br>  - "Счёт выставлен: Номер". Было {old_value} стало {new_value}"'
                    invoice.name = value

                if 'price' in invoice_args:
                    value = invoice_args['price']
                    log_with_changes += f'<br>  - "Счёт выставлен: Сумма". Было {invoice.price} стало {value}"'
                    invoice.price = value

                if 'date' in invoice_args:
                    value = invoice_args['date']
                    old_value = invoice.date.strftime(human_date) if invoice.date else 'пустым'
                    new_value = datetime.strptime(value, '%Y-%m-%d').strftime(human_date) if value else  'пустым'
                    log_with_changes += f'<br>  - "Счёт выставлен: Дата". Было {old_value} стало {new_value}"'
                    invoice.date = value

                if 'payment_expected_date' in invoice_args:
                    value = invoice_args['payment_expected_date']
                    old_value = invoice.payment_expected_date.strftime(human_date)
                    new_value = datetime.strptime(value, '%Y-%m-%d').strftime(human_date)
                    log_with_changes += f'<br>  - "Ожидаемая дата платежа". Было {old_value} стало {new_value}"'
                    invoice.payment_expected_date = value
                    
                if 'payment_amount' in invoice_args:
                    value = invoice_args['payment_amount']
                    old_value = invoice.payment_amount or 'пустым'
                    new_value = value or 'пустым'
                    log_with_changes += f'<br>  - "Платеж получен/отправлен: Сумма". Было {old_value} стало {new_value}"'
                    invoice.payment_amount = value

                if 'payment_date' in invoice_args:
                    value = invoice_args['payment_date']
                    old_value = invoice.payment_date.strftime(human_date) if invoice.payment_date else 'пустым'
                    new_value = datetime.strptime(value, '%Y-%m-%d').strftime(human_date) if value else  'пустым'
                    log_with_changes += f'<br>  - "Платеж получен/отправлен: Дата". Было {old_value} стало {new_value}"'
                    invoice.payment_date = value


                invoice.contract_id = contract_id
                existing_invoices_in_args.append(invoice.id)
                sess.commit()
            else:
                new_invoice =  Invoice(**invoice_args)
                new_invoice.contract_id = contract_id
                sess.add(new_invoice)
                log_with_changes += '<br>  - Добавлен новый счёт'
                sess.commit()
                existing_invoices_in_args.append(new_invoice.id)
        
        for existing_invoice in contract.fk__invoice:
            if existing_invoice.id not in existing_invoices_in_args:
                log_with_changes += '<br>  - счёт удалён'
                sess.delete(existing_invoice)
        sess.commit()
        
        if '<br>' in log_with_changes:
            __insert_log(user, 'Доход/Расход', log_with_changes)


        for new_link_contract_args in linked_contracts_args:
            if new_link_contract_args.get('id'):
                update_contract(new_link_contract_args['id'], new_link_contract_args)
            else:
                new_link_contract_id, _ = insert_contract(new_link_contract_args, user)
                if contract.it_is == {'Доход'}:
                    debit_credit = DebitCredit(debit=contract_id, credit=new_link_contract_id)
                else:
                    debit_credit = DebitCredit(debit=new_link_contract_id, credit=contract_id)
                sess.add(debit_credit)
        sess.commit()
        
    except Exception as e:
        sess.close()
        raise e
    sess.close()
    return None, 200

def delete_contract(contract_id, user):
    sess = Session()
    try:
        contract: Contract = sess.get(Contract, contract_id)
        contract.is_active = False
        
        if contract.it_is == {'Доход'}:
            debit_credit_list = contract.fk__debit
        else:
            debit_credit_list = contract.fk__credit
        
        for row in debit_credit_list:
            sess.delete(row)
        sess.commit()
    except Exception as e:
        sess.close()
        raise e
    sess.close()

    __insert_log(user, "Доход/Расход", f"Удален контракт:<br>  ID: {contract_id}")
    return 200


def get_relationship(contract_id: int) -> Tuple[dict, int]:
    """Получение списка противоположных контрактов с указанием свой/чужой"""
    sess = Session()
    try:
        contract: Contract = sess.get(Contract, contract_id)
        if contract is None:
            return {"message": f"Запись с идентификатором {contract_id} не найдена.<br>Обновите страницу или обратитесь к администратору системы."}, 404
        linked_contracts: List[dict] = list()
        linked_contracts_id : List[int] = list()
        if contract.it_is == {'Доход'}:
            find_type = 'Расход'
            debit_credit_list: List[DebitCredit] = contract.fk__debit
            for debit_credit in debit_credit_list:
                linked_contracts.append(debit_credit.fk__credit.to_dict())
                linked_contracts_id.append(debit_credit.fk__credit.id)
        else:
            find_type = 'Доход'
            debit_credit_list: List[DebitCredit] = contract.fk__credit
            for debit_credit in debit_credit_list:
                linked_contracts.append(debit_credit.fk__debit.to_dict())
                linked_contracts_id.append(debit_credit.fk__debit.id)
        other_contracts: List[Contract] = sess.query(Contract).filter(Contract.it_is == find_type, Contract.is_active)
        if linked_contracts_id:
            other_contracts = other_contracts.filter(Contract.id.not_in(linked_contracts_id))
        result = {
            "linked_contracts": linked_contracts,
            "other_contracts": [_.to_dict() for _ in other_contracts]
        }
        sess.close()
    except:
        sess.close()
        raise
    return result, 200


def change_relationship(contract_id: int | str, linked_contract: List[int | str]):
    """Изменение связок существующих доходов и расходов"""
    sess = Session()
    try:
        main_contract: Contract = sess.get(Contract, contract_id)
        if main_contract.it_is == {'Доход'}:
            debit_credit_list: List[DebitCredit] = main_contract.fk__debit
            for row in debit_credit_list:
                if row.credit in linked_contract:
                    linked_contract.pop(linked_contract.index(row.credit))
                else:
                    sess.delete(row)
            add_relationship = list()
            for credit_id in linked_contract:
                credit: Contract = sess.get(Contract, credit_id)
                if credit is not None and credit.it_is == {'Расход'}:
                    add_relationship.append(
                        DebitCredit(
                            debit=contract_id, 
                            credit=credit_id
                        )
                    )

        else:
            debit_credit_list: List[DebitCredit] = main_contract.fk__credit
            for row in debit_credit_list:
                if row.debit in linked_contract:
                    linked_contract.pop(linked_contract.index(row.debit))
                else:
                    sess.delete(row)
            add_relationship = list()
            for debit_id in linked_contract:
                debit: Contract = sess.get(Contract, debit_id)
                if debit is not None and debit.it_is == {'Доход'}:
                    add_relationship.append(
                        DebitCredit(
                            debit=debit_id, 
                            credit=contract_id
                        )
                    )
        if add_relationship:
            sess.add_all(add_relationship)
        sess.commit()
        sess.close()
    except Exception as e:
        sess.close()
        return 500
    return 200


def get_account():

    def get_plan(it_is: Literal['Доход', 'Расход']):
        start_date = date.today().replace(day=1)
        end_date = (np.datetime64(start_date.strftime('%Y-%m')) + np.timedelta64('1', 'M')).astype(date)
        result = 0
        contract_id_list = list()
        plan_contract: List[Contract] = (
            sess.query(Contract)
            .join(Invoice, isouter=True)
            .filter(
                or_(
                    and_(
                        Contract.order_deadline >= start_date,
                        Contract.order_deadline < end_date,
                    ),
                    and_(
                        Contract.repeating_month >= start_date,
                        Contract.repeating_month < end_date,
                    ),
                    and_(
                        Invoice.payment_expected_date >= start_date,
                        Invoice.payment_expected_date < end_date,
                    ),
                    and_(
                        Invoice.payment_date >= start_date,
                        Invoice.payment_date < end_date,
                    ),
                ),
                Contract.it_is == it_is,
                Contract.is_active,
            )
        )
        for contract in plan_contract:
            contract_id_list.append(contract.id)
            contract_price = contract.order_price
            invoice: Invoice
            for invoice in contract.fk__invoice:
                payment_expected_date_in_range = start_date <= invoice.payment_expected_date < end_date
                
                if not (invoice.payment_amount is None and payment_expected_date_in_range):
                    contract_price -= (invoice.price)
            result += contract_price if contract_price > 0 else 0
        return result

    result = dict()

    sess = Session()
    try:
        account: PaymentAccount = (
            sess.query(PaymentAccount)
            .filter(PaymentAccount.is_active)
            .order_by(PaymentAccount.start_date.desc())
            .limit(1)
            .one()
        )
        account_balance = account.Money
        account_date = account.start_date

        debit_on_current_date = (
            sess.query(func.sum(Invoice.payment_amount))
            .join(Contract)
            .filter(
                Contract.it_is == 'Доход',
                Contract.is_active,
                Invoice.payment_date < date.today(),
                Invoice.payment_date >= account_date,
            )
            .scalar()
        ) or 0

        credit_on_current_date = (
            sess.query(func.sum(Invoice.payment_amount))
            .join(Contract)
            .filter(
                Contract.it_is == 'Расход',
                Contract.is_active,
                Invoice.payment_date < date.today(),
                Invoice.payment_date >= account_date,
            )
            .scalar()
        ) or 0
        
        current_balance = account_balance + debit_on_current_date - credit_on_current_date
        result['fact'] = current_balance

        debit = get_plan('Доход')
        credit = get_plan('Расход')

        result['plan'] = current_balance + debit - credit

        sess.close()
    except Exception as e:
        sess.close()
        return None, 500
    return result, 200


def get_dashboard_debit_credit(start: str = None, end: str = None):
    """Получение данных дохода/расхода по текущему году"""
    current_date = (date.today()).replace(day=1)
    current_date_str = current_date.strftime("%Y-%m")
    weights = 3 
    if not (start or end):
        start = current_date.strftime("%Y-01")
        end = current_date.strftime("%Y-12")
    try:
        np_start_date = np.datetime64(start)
        np_end_date = np.datetime64(end) + np.timedelta64('1', 'M')
        start_date: date = np_start_date.astype(date)
        end_date: date = (np_end_date + np.timedelta64('2', 'M')).astype(date) # + 2 месяца для построения линии тренда
        date_to_remove: List[date] = np.arange(np_end_date, np_end_date + np.timedelta64('2', 'M'), np.timedelta64('1', 'M')).astype(date)
        months_all = {_.strftime("%Y-%m"): 0 for _ in np.arange(np_start_date, np_end_date + np.timedelta64('2', 'M'), np.timedelta64('1', 'M')).astype(date)}
    except:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400
    months_fact = {_: 0 for _ in months_all if _ <= current_date_str}
    months_plan = {_: 0 for _ in months_all if _ >= current_date_str}

    def get_fact(it_is: Literal['Доход', 'Расход']):
        current_end_date = np.datetime64(current_date_str) + np.timedelta64('1', 'M')
        _end_date = current_end_date.astype(date) if current_end_date < np_end_date else end_date
        result = months_fact.copy()
        if start_date > current_date:
            return result
        
        fact = (
            sess.query(func.year(Invoice.payment_date), func.month(Invoice.payment_date), func.sum(Invoice.payment_amount))
            .join(Contract)
            .filter(
                Contract.it_is == it_is,
                Contract.is_active,
                Invoice.payment_date < _end_date,
                Invoice.payment_date >= start_date,
            )
            .group_by(func.year(Invoice.payment_date), func.month(Invoice.payment_date))
        )
        result.update({f'{year}-{str(month).zfill(2)}': value for year, month, value in fact})
        return result
    
    def get_plan(it_is: Literal['Доход', 'Расход']):
        _start_date = current_date if current_date > start_date else start_date
        result = months_plan.copy()
        if end_date < current_date:
            return result
        plan_contract: List[Contract] = (
            sess.query(Contract)
            .join(Invoice, isouter=True)
            .filter(
                or_(
                    and_(
                        Contract.order_deadline >= _start_date,
                        Contract.order_deadline < end_date,
                    ),
                    and_(
                        Contract.repeating_month >= _start_date,
                        Contract.repeating_month < end_date,
                    ),
                    and_(
                        Invoice.payment_expected_date >= start_date,
                        Invoice.payment_expected_date < end_date,
                    ),
                ),
                Contract.it_is == it_is,
                Contract.is_active,
            )
        )
        for contract in plan_contract:
            contract_price = contract.order_price
            invoice: Invoice
            if contract.repeating_month is None:
                _month = contract.order_deadline
            elif contract.repeating_month >= _start_date and contract.repeating_month < end_date: 
                _month = contract.repeating_month
            else:
                continue
            for invoice in contract.fk__invoice:
                contract_price -= invoice.price
                if invoice.payment_expected_date >= (current_date if current_date > start_date else start_date) and invoice.payment_expected_date < end_date:
                    result[invoice.payment_expected_date.strftime("%Y-%m")] += invoice.price
            if _start_date <= _month < end_date:
                result[_month.strftime("%Y-%m")] += contract_price if contract_price > 0 else 0
        for k, v in result.items():
            result[k] = round(v,2)
        return result


    sess = Session()
    try:
        debit_fact: Dict[str, int] = get_fact('Доход')
        credit_fact: Dict[str, int] = get_fact('Расход')
        debit_plan: Dict[str, int] = get_plan('Доход')
        credit_plan: Dict[str, int] = get_plan('Расход')
        sess.close()
    except:
        sess.close()
        raise
    
    debit_all = debit_fact.copy()
    debit_all.update(debit_plan)
    debit_trend_value = np.convolve(list(debit_all.values()), np.ones(weights), 'valid')/weights
    
    credit_all = credit_fact.copy()
    credit_all.update(credit_plan)
    credit_trend_value = np.convolve(list(credit_all.values()), np.ones(weights), 'valid')/weights
    
    for d in date_to_remove:
        key = d.strftime("%Y-%m")
        if key in debit_plan:
            debit_plan.pop(key)
            credit_plan.pop(key)

    result = {
        'debit': {
            'fact': debit_fact,
            'plan': debit_plan,
            'trend': {k:round(v, 2) for k, v in zip(debit_all, debit_trend_value)},
        },
        'credit': {
            'fact': credit_fact,
            'plan': credit_plan,
            'trend': {k:round(v, 2) for k, v in zip(credit_all, credit_trend_value)},
        },
        'delta': {
            'fact': {key: round(debit_fact[key] - credit_fact[key], 2)  for key in months_fact},
            'plan': {key: round(debit_plan[key] - credit_plan[key], 2)  for key in debit_plan},
            'trend': {key: round(debit - credit, 2) for key, debit, credit in zip(debit_all, debit_trend_value, credit_trend_value)},
        },
    }
    return result, 200


def export_contracts(query_token):
    
    sess = Session()
    try:
        contracts = query_cache.get(query_token)
        if contracts is None:
            raise my_exc.NotContracts

        wb = Workbook()
        ws = wb.active
        # Шапка

        head = [
            (
                'ID',
                'тип',
                'ежемесячный',
                'контрагент',
                'Тип дохода/расхода',
                'Описание работ/продукции',
                'Номер договора',
                'Дата договора',
                'Номер заказа',
                'Дата заказа',
                'Общая стоимость заказа, с НДС',
                'Дата окончания заказа',
                'Счет выставлен',
                '',
                'Ожидаемая дата платежа',
                'Платеж получен/отправлен',
                '',
            ),
            (
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                'Сумма, с НДС',
                'Дата',
                '',
                'Сумма, с НДС',
                'Дата',
            ),
        ]
        ws.append(head[0])
        ws.append(head[1])
        for row in contracts:
            for i in row.to_export():
                ws.append(i)
        
        sess.close()
        numbers.FORMAT_CURRENCY_USD
    except my_exc.NotContracts:
        sess.close()
        return {'message': 'Запрос устарел. Обновите страницу'}, 400
    except Exception as e:
        sess.close()
        return {'message': 'непредвиденная ошибка'}, 500
    
    # настройка визуальной части
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')
    ws.merge_cells('H1:H2')
    ws.merge_cells('I1:I2')
    ws.merge_cells('J1:J2')
    ws.merge_cells('K1:K2')
    ws.merge_cells('L1:L2')
    ws.merge_cells('M1:N1')
    ws.merge_cells('O1:O2')
    ws.merge_cells('P1:R1')

    for cell in ws['C'][2:] + ws['H'][2:] + ws['J'][2:] + ws['K'][2:] + ws['N'][2:] + ws['O'][2:] + ws['R'][2:]:
        cell.number_format = 'dd.mm.yyyy'

    for cell in ws['P'][2:] + ws['M'][2:] + ws['K'][2:]:
        cell.number_format = '### ### ### ### ### ##0.00 ₽'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 27
    ws.column_dimensions['G'].width = 21
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 21
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 14
    ws.column_dimensions['R'].width = 15

    bytes_book = BytesIO()
    wb.save(bytes_book)
    wb.close()
    bytes_book.seek(0)

    return bytes_book
