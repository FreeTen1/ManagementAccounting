from typing import Any, Dict, List, Literal, Tuple
from db_models import (
    PaymentAccount,
    Contractor,
    ContractType,
    RepeatingContract,
    Contract,
    Invoice,
    DebitCredit
)
import numpy as np
from my_engine import Session
from sqlalchemy import and_, func, or_, text
from sqlalchemy.orm.query import Query
from datetime import date
import my_exceptions as my_exc
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO


TWO_SPACE = '&nbsp;&nbsp;'
FOUR_SPACE = '&nbsp;&nbsp;&nbsp;&nbsp;'

def _total_by_contractors(start_date: date, end_date: date, contractors: List[str] = None, it_is = 'Доход'):

    sess = Session()
    try:
        query = (
            sess.query(
                Contractor.id, 
                Contractor.name, 
                Contractor.color_hex,
                func.sum(Invoice.payment_amount).label('val')
            )
            .select_from(Contract)
            .join(Invoice)
            .join(Contractor)
            .filter(
                Contract.it_is == it_is,
                Contract.is_active,
                Contractor.is_active,
                Invoice.payment_date >= start_date,
                Invoice.payment_date < end_date,
            )
            .group_by(Contractor.id, Contractor.name, Contractor.color_hex)
            .order_by(text('val DESC'))
        )
        if contractors:
            query = query.filter(Contractor.id.in_(contractors))
        sum_by_contractors = {
            _[0] : {
                "id": _[0],
                'name': _[1],
                'color': _[2],
                'value': _[-1],
            }
            for _ in query
        }
    except Exception as e:
        return None
    finally:
        sess.close()
    
    return sum_by_contractors


def _total_plan_by_contractors(start_date: date, end_date: date, contractors: List[str] = None, it_is = 'Доход'):

    sess = Session()
    try:
        result = dict()

        plan_contract: Query = (
            sess.query(Contract)
            .join(Invoice, isouter=True)
            .filter(
                or_(
                    and_(
                        Contract.order_deadline >= start_date,
                        Contract.order_deadline < end_date,
                    ),
                    and_(
                        Contract.repeating_month >= start_date,
                        Contract.repeating_month < end_date,
                    ),
                    and_(
                        Invoice.payment_expected_date >= start_date,
                        Invoice.payment_expected_date < end_date,
                    ),
                    and_(
                        Invoice.payment_date >= start_date,
                        Invoice.payment_date < end_date,
                    ),
                ),
                Contract.it_is == it_is,
                Contract.is_active,
            )
        )

        if contractors:
            plan_contract = plan_contract.filter(
                Contractor.id.in_(contractors)
            )
        contract: Contract
        for contract in plan_contract:
            contract_price = contract.order_price
            contractor_id = contract.contractor_id
            main_date = contract.repeating_month or contract.order_deadline
            if contractor_id not in result:
                result[contractor_id] = {
                    "id": contractor_id,
                    'name': contract.fk__contractor.name,
                    'color': contract.fk__contractor.color_hex,
                    'value': 0,
                }
            
            invoice: Invoice
            for invoice in contract.fk__invoice:
                if invoice.payment_date:
                    contract_price -= invoice.payment_amount
                    if start_date <= invoice.payment_date < end_date:
                        result[contractor_id]['value'] += invoice.payment_amount
                elif invoice.payment_expected_date:
                    contract_price -= invoice.price
                    if start_date <= invoice.payment_expected_date < end_date:
                        result[contractor_id]['value'] += invoice.price

            if contract_price > 0 and start_date <= main_date < end_date:
                result[contractor_id]['value'] += contract_price
            elif result[contractor_id]['value'] == 0:
                result.pop(contractor_id)
    except Exception as e:
        return None
    finally:
        sess.close()
    
    for v in result.values():
        v['value'] = round(v['value'], 2)

    return result


def _total_income_by_years(contractor = None):
    current_year = date.today().year
    sess = Session()
    try:
        query = (
            sess.query(func.year(Invoice.payment_date) ,func.sum(Invoice.payment_amount))
            .select_from(Contract)
            .join(Invoice)
            .filter(
                Contract.it_is == 'Доход',
                Contract.is_active,
                Invoice.payment_date >= f"{current_year-3}-01-01",
                Invoice.payment_date <= f"{current_year}-12-31",
            )
            .group_by(text('year_1'))
            .order_by(text('year_1 DESC'))
        )
        if contractor:
            query = query.filter(Contract.contractor_id == contractor)
        result = [{'year': row[0], 'value': round(row[1], 2)} for row in query]
    except Exception as e:
        return None
    finally:
        sess.close()
    return result


def _total_plan_income_by_years(contractor = None, it_is = 'Доход'):

    current_year = date.today().year
    sum_by_years = {_ : {'year': _, 'value': 0} for _ in range(current_year-3, current_year+1)}
    sess = Session()
    try:
        plan_contract: Query = (
            sess.query(Contract)
            .join(Invoice, isouter=True)
            .filter(
                or_(
                    and_(
                        Contract.order_deadline >= f"{current_year-3}-01-01",
                        Contract.order_deadline <= f"{current_year}-12-31",
                    ),
                    and_(
                        Contract.repeating_month >= f"{current_year-3}-01-01",
                        Contract.repeating_month <= f"{current_year}-12-31",
                    ),
                    and_(
                        Invoice.payment_expected_date >= f"{current_year-3}-01-01",
                        Invoice.payment_expected_date <= f"{current_year}-12-31",
                    ),
                    and_(
                        Invoice.payment_date >= f"{current_year-3}-01-01",
                        Invoice.payment_date <= f"{current_year}-12-31",
                    ),
                ),
                Contract.it_is == it_is,
                Contract.is_active,
            )
        )

        if contractor:
            plan_contract = plan_contract.filter(Contract.contractor_id == contractor)

        for contract in plan_contract:
            contract: Contract
            contract_price = contract.order_price
            invoice: Invoice
            for invoice in contract.fk__invoice:
                if invoice.payment_date:
                    contract_price -= invoice.payment_amount
                    if invoice.payment_date.year in sum_by_years:
                        sum_by_years[invoice.payment_date.year]['value'] += invoice.payment_amount

                elif invoice.payment_expected_date:
                    contract_price -= invoice.price
                    if invoice.payment_expected_date.year in sum_by_years:
                        sum_by_years[invoice.payment_expected_date.year]['value'] += invoice.price
            
            _years = contract.repeating_month.year if contract.repeating_month else contract.order_deadline.year
            if contract_price > 0 and _years in sum_by_years:
                sum_by_years[_years]['value'] += contract_price
    except Exception as e:
        return None
    finally:
        sess.close()

    result = [_ for _ in sum_by_years.values() if _['value']]
    return result


def total_income_by_years(contractor = None):

    total_plan = _total_plan_income_by_years(contractor)
    # total_fact = _total_income_by_years(contractor)
    
    # return {
    #     'plan':{
    #         'year_by_contractors': total_plan,
    #         'max_values': {'year_by_contractors': _max(total_plan)}
    #     },
    #     'fact':{
    #         'year_by_contractors': total_fact,
    #         'max_values': {'year_by_contractors': _max(total_fact)}
    #     },
    # }, 200

    return {
        'year_by_contractors': total_plan,
        'max_values': {'year_by_contractors': _max(total_plan)}
    }, 200

def _append_precent(current: dict, old: dict):

    for contractor in current:
            if data := old.get(contractor):
                percent = round((current[contractor]['value'] - data['value']) / data['value'] * 100)
                if percent > 0:
                    current[contractor]['percent'] = percent
                    current[contractor]['is_increased'] = True
                elif percent == 0:
                    current[contractor]['percent'] = percent
                    current[contractor]['is_increased'] = None
                else:
                    current[contractor]['percent'] = -percent
                    current[contractor]['is_increased'] = False
            else:
                current[contractor]['percent'] = 100
                current[contractor]['is_increased'] = True
    

def _max(iter, key = 'value') -> int:
    if not iter:
        return 0
    if isinstance(iter, dict):
        result = max(iter.values(), key=lambda x:x[key])[key]
    else:
        result = max(iter, key=lambda x:x[key])[key]
    return round(result, 2)


def _total_income(start_date, end_date, plan_or_fact, add_percent_sum):
    assert plan_or_fact in ["plan", "fact"], "_total_income: plan_or_fact wrong value"

    if plan_or_fact == "plan":
        _func = _total_plan_by_contractors
        # _func_year = _total_plan_income_by_years
    else:
        _func = _total_by_contractors
        # _func_year = _total_income_by_years

    result = dict()
    # max_values = dict()
    # result['max_values'] = max_values

    sum_by_contractors_peroid = _func(start_date.astype(date), end_date.astype(date))
    if add_percent_sum:
        sum_for_percent = _func(
            (start_date - np.timedelta64('1', 'M')).astype(date), 
            (end_date - np.timedelta64('1', 'M')).astype(date), 
            sum_by_contractors_peroid.keys()
        )
        _append_precent(sum_by_contractors_peroid, sum_for_percent)
    
    # max_values['period_by_contractors'] = _max(sum_by_contractors_peroid)

    result['period_by_contractors'] = list(sum_by_contractors_peroid.values())
    

    # by_years = _func_year()
    # max_values['by_years'] = _max(by_years)

    # result['by_years'] = by_years
    
    return result


def total_income(start_date = None, end_date = None):

    result = dict()
    max_values = dict()
    result['max_values'] = max_values
    current_date = date.today().replace(day=1)
    try:
        if start_date is None and end_date is None:
            start_date = np.datetime64(current_date.strftime("%Y-%m"))
            end_date = start_date + np.timedelta64('1', 'M')
            add_percent_sum = True 
        elif start_date is None or end_date is None:
            return {'message': f'ошибка в датах. Нет даты старта или конца.\n start_date={bool(start_date)}, end_date={bool(end_date)}'}, 400 
        elif start_date > end_date:
            return {'message': 'ошибка в датах. Дата старта больше даты конца'}, 400
        elif start_date and end_date:
            start_date = np.datetime64(start_date)
            end_date = np.datetime64(end_date) + np.timedelta64('1', 'M')
            add_percent_sum = False 
    except Exception:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400

    total_income_plan = _total_income(start_date, end_date, "plan", add_percent_sum)
    total_income_fact = _total_income(start_date, end_date, "fact", add_percent_sum)

    contractor: dict
    fact_contractor: dict
    for contractor in total_income_plan['period_by_contractors']:
        contractor_id = contractor['id']
        plan = {
            'value': contractor.pop('value')
        }
        contractor['plan'] = plan

        if 'percent' in contractor:
            plan['percent'] = contractor.pop('percent')
            plan['is_increased'] = contractor.pop('is_increased')

        for fact_contractor in total_income_fact['period_by_contractors']:
            if contractor_id == fact_contractor['id']:
                fact = {
                    'value': fact_contractor['value']
                }
                contractor['fact'] = fact

                if 'percent' in fact_contractor:
                    fact['percent'] = fact_contractor['percent']
                    fact['is_increased'] = fact_contractor['is_increased']
                break
        else:
            contractor['fact'] = {
                'value': 0
            }
    
    result['period_by_contractors'] = total_income_plan['period_by_contractors']

    sum_by_contractors_year = _total_plan_by_contractors(
        current_date.replace(month=1), 
        current_date.replace(year=current_date.year+1, month=1)
    )
    sum_for_percent = _total_plan_by_contractors(
        current_date.replace(year=current_date.year-1, month=1),
        current_date.replace(month=1),
        sum_by_contractors_year.keys()
    )
    _append_precent(sum_by_contractors_year, sum_for_percent)

    max_values['year_by_contractors'] = _max(sum_by_contractors_year)
    result['year_by_contractors'] = list(sum_by_contractors_year.values())

    by_years = _total_plan_income_by_years()
    max_values['by_years'] = _max(by_years)

    result['by_years'] = by_years


    result['info'] = {
        'by_years': 'Плановый доход за текущий год по сравнению с предыдущими годами.<br>'
                    'Считается как плановый доход в \"Доход от контрагентов\".',
        'year_by_contractors': 'Плановый доход за текущий год с разбивкой по контрагентам.<br>'
                               'Считается как плановый доход в \"Доход от контрагентов\".',
        'period_by_contractors': 'Плановый и фактический доход за выбранный промежуток (по умолчанию за текущий месяц)<br>'
                                 'Фактический доход считается по блоку \"Платёж получен\":<br>'
                                 f'{TWO_SPACE}1. Ищет даты, которые попадают в выбранный диапазон<br>'
                                 f'{TWO_SPACE}2. Суммирует найденные счета<br>'
                                 '<br>'
                                 'Плановый доход:<br>'
                                 f'{TWO_SPACE}1. Ищет все контракты дохода, у которых дата попадает в выбранный диапазон в следующих столбцах:<br>'
                                 f'{FOUR_SPACE}а. Платёж получен: Дата<br>'
                                 f'{FOUR_SPACE}б. Ожидаемая дата платежа<br>'
                                 f'{FOUR_SPACE}в. Месяц платежа<br>'
                                 f'{FOUR_SPACE}г. Дата окончания заказа<br>'
                                 f'{TWO_SPACE}2. Если \"Месяц платежа\" или \"Дата окончания заказа\" входят в диапазон<br>'
                                 f'{FOUR_SPACE}а. Общая стоимость заказа добавляется к общей сумме<br>'
                                 f'{FOUR_SPACE}б. Счета ВНЕ выбранного диапазона вычитаются из общей суммы<br>'
                                 f'{TWO_SPACE}3. Если \"Месяц платежа\" или \"Дата окончания заказа\" ВНЕ выбранного диапазона<br>'
                                 f'{FOUR_SPACE}а. К общей сумме прибавляю только счета в выбранном диапазоне<br>'
                                 '<br>'
                                 'Правила вычитания счетов:<br>'
                                 f'{TWO_SPACE}1. Если дата в столбце \"Платёж получен: Дата\" ВНЕ выбранного диапазона, то вычитается \"Платёж получен: Сумма\"<br>'
                                 f'{TWO_SPACE}2. Если столбец \"Платёж получен: Дата\" пуст, а дата в столбце \"Ожидаемая дата платежа\" ВНЕ выбранного диапазона, то вычитается \"Счёт выставлен: Сумма\"<br>'
                                 f'{TWO_SPACE}3. Все остальные счета игнорируются<br>'
                                 '<br>'
                                 'Правила прибавления счетов:<br>'
                                 f'{TWO_SPACE}1. Если дата в столбце \"Платёж получен: Дата\" в выбранном диапазоне, то прибавляется \"Платёж получен: Сумма\"<br>'
                                 f'{TWO_SPACE}2. Если столбец в столбце \"Платёж получен: Дата\" пуст, а дата в столбце \"Ожидаемая дата платежа\" в выбранном диапазоне, то прибавляется \"Счёт выставлен: Сумма\"<br>'
                                 f'{TWO_SPACE}3. Все остальные счета игнорируются<br>',
    }

    return result, 200


def _by_contract_type(start_date: date, end_date: date, contractor: int, it_is = 'Доход'):

    sess = Session()
    try:
        query = (
            sess.query(ContractType.id, ContractType.name, func.sum(Invoice.payment_amount).label('val'))
            .select_from(Contract)
            .join(Invoice)
            .join(ContractType)
            .filter(
                Contract.contractor_id == contractor,
                Contract.is_active,
                Contract.it_is == it_is,
                Invoice.payment_date >= start_date,
                Invoice.payment_date < end_date,
            )
            .group_by(ContractType.id, ContractType.name)
            .order_by(text('val DESC'))
        )
        sum_by_contract_types = {
            _[0] : {
                "id": _[0],
                'name': _[1],
                'value': _[2],
            }
            for _ in query
        }
    except Exception as e:
        return None
    finally:
        sess.close()
    
    return sum_by_contract_types


def income_by_contract_type(start_date = None, end_date = None):

    result = dict()
    max_values = dict()
    result['max_values'] = max_values
    current_date = date.today().replace(day=1)

    try:
        if start_date is None and end_date is None:
            start_date = np.datetime64(current_date.strftime("%Y-%m"))
            end_date = start_date + np.timedelta64('1', 'M')
            add_percent_sum = True 
        elif start_date is None or end_date is None:
            return {'message': f'ошибка в датах. Нет даты старта или конца.\n start_date={bool(start_date)}, end_date={bool(end_date)}'}, 400 
        elif start_date > end_date:
            return {'message': 'ошибка в датах. Дата старта больше даты конца'}, 400
        elif start_date and end_date:
            start_date = np.datetime64(start_date)
            end_date = np.datetime64(end_date) + np.timedelta64('1', 'M')
            add_percent_sum = False 
    except Exception:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400
    
    sum_by_contractors = _total_by_contractors(start_date.astype(date), end_date.astype(date))
    if not sum_by_contractors:
        return {"message": "За выбранный промежуток нет данных"}, 404


    total_sum_by_contract_types = dict()
    total_sum = 0
    for contractor_id, data in sum_by_contractors.items():
        sum_by_contract_types = _by_contract_type(start_date.astype(date), end_date.astype(date), contractor_id)
        if add_percent_sum:
            sum_for_percent = _by_contract_type(
                (start_date - np.timedelta64('1', 'M')).astype(date), 
                (end_date - np.timedelta64('1', 'M')).astype(date), 
                contractor_id
            )
            _append_precent(sum_by_contract_types, sum_for_percent)
        data['contract_types'] = list(sum_by_contract_types.values())
        for contract_type_id, sum_data in sum_by_contract_types.items():
            if total_sum_by_contract_types.get(contract_type_id):
                total_sum_by_contract_types[contract_type_id]['value'] += sum_data['value']
                total_sum_by_contract_types[contract_type_id]['contractors'].append(contractor_id)
            else:
                total_sum_by_contract_types[contract_type_id] = {
                    'id': sum_data['id'],
                    'name': sum_data['name'],
                    'value': sum_data['value'],
                    'contractors': [contractor_id],
                }
            total_sum += sum_data['value']
    
    result['by_contractors'] = list(sum_by_contractors.values())
    result['by_types'] = list(total_sum_by_contract_types.values())
    result['total_sum'] = round(total_sum, 2)

    result['info'] = {
        'by_contractors': 'Фактический доход за выбранный период с разбивкой по контрагентам.'
                          ' Дополнительно контрагент разбивается на типы контрактов.<br>'
                          'Фактический доход считается по блоку \"Платёж получен\":<br>'
                          f'{TWO_SPACE}1. Ищет даты, которые попадают в выбранный диапазон<br>'
                          f'{TWO_SPACE}2. Суммирует найденные счета<br>',
        'by_types': 'Фактический доход за выбранный период с разбивкой по типам контрактов.<br>'
    }

    max_values['by_contractors'] = _max(sum_by_contractors)
    max_values['by_types'] = _max(total_sum_by_contract_types)

    return result, 200


def expenses_by_contract_type(start_date = None, end_date = None):

    result = dict()
    max_values = dict()
    result['max_values'] = max_values
    current_date = date.today().replace(day=1)

    try:
        if start_date is None and end_date is None:
            start_date = np.datetime64(current_date.strftime("%Y-%m"))
            end_date = start_date + np.timedelta64('1', 'M')
        elif start_date is None or end_date is None:
            return {'message': f'ошибка в датах. Нет даты старта или конца.\n start_date={bool(start_date)}, end_date={bool(end_date)}'}, 400 
        elif start_date > end_date:
            return {'message': 'ошибка в датах. Дата старта больше даты конца'}, 400
        elif start_date and end_date:
            start_date = np.datetime64(start_date)
            end_date = np.datetime64(end_date) + np.timedelta64('1', 'M')
    except Exception:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400
    
    sum_by_contractors = _total_by_contractors(start_date.astype(date), end_date.astype(date), it_is='Расход')
    if not sum_by_contractors:
        return {"message": "За выбранный промежуток нет данных"}, 404

    total_sum_by_contract_types = dict()
    total_sum = 0
    for contractor_id, data in sum_by_contractors.items():
        sum_by_contract_types = _by_contract_type(start_date.astype(date), end_date.astype(date), contractor_id, 'Расход')

        data['contract_types'] = list(sum_by_contract_types.values())
        for contract_type_id, sum_data in sum_by_contract_types.items():
            if total_sum_by_contract_types.get(contract_type_id):
                total_sum_by_contract_types[contract_type_id]['value'] += sum_data['value']
                total_sum_by_contract_types[contract_type_id]['contractors'].append(contractor_id)
            else:
                total_sum_by_contract_types[contract_type_id] = {
                    'id': sum_data['id'],
                    'name': sum_data['name'],
                    'value': sum_data['value'],
                    'contractors': [contractor_id],
                }
            total_sum += sum_data['value']
    
    result['by_contractors'] = list(sum_by_contractors.values())
    result['by_types'] = list(total_sum_by_contract_types.values())
    result['total_sum'] = round(total_sum, 2)

    result['info'] = {
        'by_contractors': 'Фактический расход за выбранный период с разбивкой по контрагентам.'
                          ' Дополнительно контрагент разбивается на типы контрактов<br>'
                          'Фактический расход считается по блоку \"Платёж получен\":<br>'
                          f'{TWO_SPACE}1. Ищет даты, которые попадают в выбранный диапазон<br>'
                          f'{TWO_SPACE}2. Суммирует найденные счета<br>',
        'by_types': 'Фактический расход за выбранный период с разбивкой по типам контрактов'
    }

    max_values['by_contractors'] = _max(sum_by_contractors)
    max_values['by_types'] = _max(total_sum_by_contract_types)

    return result, 200


def _unpaid(start_date: date, end_date: date, it_is: str):

    sess = Session()
    try:
        query = (
            sess.query(
                Contractor.id, 
                Contractor.name, 
                ContractType.id, 
                ContractType.name, 
                Contractor.color_hex, 
                func.sum(Invoice.price), 
                func.count(1)
            )
            .select_from(Contract)
            .join(Invoice)
            .join(Contractor)
            .join(ContractType)
            .filter(
                Contract.is_active,
                Contract.it_is == it_is,
                Invoice.payment_expected_date >= start_date,
                Invoice.payment_expected_date < end_date,
                Invoice.payment_date.is_(None),
            )
            .group_by(Contractor.id, Contractor.name, ContractType.id, ContractType.name, Contractor.color_hex)
            .order_by(Contractor.id, text('sum_1 DESC'))
        )
        sum_by_contractors = [
            {
                'id': _[0],
                'name': _[1],
                'color': _[4],
                'contract_types': [{
                    'id': _[2],
                    'name': _[3],
                    'value': _[-2],
                    'count': _[-1],
                }]
            } for _ in query
        ]
    except Exception as e:
        return None
    finally:
        sess.close()
    
    return sum_by_contractors


def unpaid(start_date = None, end_date = None, it_is: str = 'Доход'):

    result = {
        'by_contractors': list(),
        'by_types': list(),
    }
    max_values = dict()
    result['max_values'] = max_values

    current_date = date.today().replace(day=1)
    try:
        if start_date is None and end_date is None:
            start_date = np.datetime64(current_date.strftime("%Y-%m"))
            end_date = start_date + np.timedelta64('1', 'M')
        elif start_date is None or end_date is None:
            return {'message': f'ошибка в датах. Нет даты старта или конца.\n start_date={bool(start_date)}, end_date={bool(end_date)}'}, 400 
        elif start_date > end_date:
            return {'message': 'ошибка в датах. Дата старта больше даты конца'}, 400
        elif start_date and end_date:
            start_date = np.datetime64(start_date)
            end_date = np.datetime64(end_date) + np.timedelta64('1', 'M')
    except Exception:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400
    
    all_unpaid_invoice = _unpaid(start_date.astype(date), end_date.astype(date), it_is)
    if not all_unpaid_invoice:
        return {"message": "За выбранный промежуток нет данных"}, 404

    total_sum = total_count = cur_sum = cur_count = 0
    by_types = dict()
    cur_contractor = {'id': None}
    for contractor_data in all_unpaid_invoice:
        if contractor_data['id'] != cur_contractor['id']:
            cur_contractor['value'] = cur_sum
            cur_contractor['count'] = cur_count
            if cur_contractor['id'] is not None:
                result['by_contractors'].append(cur_contractor)
            cur_contractor = contractor_data
            total_sum += cur_sum
            total_count += cur_count
            cur_sum = cur_count = 0
        else:
            cur_contractor['contract_types'].extend(contractor_data['contract_types'])

        contract_type_data = contractor_data['contract_types'][0]
        cur_sum += contract_type_data['value']
        cur_count += contract_type_data['count']

        if contract_type_data['id'] in by_types:
            by_types[contract_type_data['id']]['value'] += contract_type_data['value']
            by_types[contract_type_data['id']]['count'] += contract_type_data['count']
            if cur_contractor['id'] not in by_types[contract_type_data['id']]['contractors']:
                by_types[contract_type_data['id']]['contractors'].append(cur_contractor['id'])
        else:
            by_types[contract_type_data['id']] = contract_type_data.copy()
            by_types[contract_type_data['id']]['contractors'] = [cur_contractor['id']]
    else:
        cur_contractor['value'] = cur_sum
        cur_contractor['count'] = cur_count
        if cur_contractor['id'] is not None:
            result['by_contractors'].append(cur_contractor)
        total_sum += cur_sum
        total_count += cur_count
    
    result['by_types'] = list(by_types.values())
    result['total_sum'] = round(total_sum, 2)
    result['total_count'] = total_count

    result['info'] = {
        'by_contractors': f'Количество и сумма счетов у контрактов \"{it_is}\" без заполненного блока \"Платёж получен\" за выбранный диапазон.<br>'
                           'Группировка по контрагентам. Дополнительно контрагент разбивается на типы контрактов',
        'by_types': f'Количество и сумма счетов у контрактов \"{it_is}\" без заполненного блока \"Платёж получен\" за выбранный диапазон.<br>'
                     'Группировка по типам контрактов.<br>'
                     'При нажатии на тип контракта в левой части останутся контрагенты только с выбранным типом'
    }

    max_values['by_contractors'] = _max(result['by_contractors'])
    max_values['by_types'] = _max(by_types)

    return result, 200


def balance_forecast(end: str = None):
    """Получение баланса на счёте на начало месяца"""

    def get_fact(it_is: Literal['Доход', 'Расход']):
        fact = (
            sess.query(func.sum(Invoice.payment_amount))
            .join(Contract)
            .filter(
                Contract.it_is == it_is,
                Contract.is_active,
                Invoice.payment_date < current_month,
                Invoice.payment_date >= account_date,
            )
        )
        return fact.scalar() or 0
    
    def get_plan(it_is: Literal['Доход', 'Расход']):
        start_date = current_month
        result = months_plan.copy()

        plan_contract: List[Contract] = (
            sess.query(Contract)
            .join(Invoice, isouter=True)
            .filter(
                or_(
                    and_(
                        Contract.order_deadline >= start_date,
                        Contract.order_deadline < end_date,
                    ),
                    and_(
                        Contract.repeating_month >= start_date,
                        Contract.repeating_month < end_date,
                    ),
                    and_(
                        Invoice.payment_expected_date >= start_date,
                        Invoice.payment_expected_date < end_date,
                    ),
                    and_(
                        Invoice.payment_date >= start_date,
                        Invoice.payment_date < end_date,
                    ),
                ),
                Contract.it_is == it_is,
                Contract.is_active,
            )
        )
        for contract in plan_contract:
            contract_price = contract.order_price
            invoice: Invoice
            if contract.repeating_month is None:
                _month = contract.order_deadline
            elif start_date <= contract.repeating_month < end_date: 
                _month = contract.repeating_month
            else:
                _month = date(1935,1,1)
            for invoice in contract.fk__invoice:
                contract_price -= invoice.price
                if invoice.payment_date and start_date <= invoice.payment_date < end_date:
                    result[invoice.payment_date.strftime("%Y-%m")] += invoice.payment_amount
                elif not invoice.payment_date and start_date <= invoice.payment_expected_date < end_date:
                    result[invoice.payment_expected_date.strftime("%Y-%m")] += invoice.price
            if start_date <= _month < end_date:
                result[_month.strftime("%Y-%m")] += contract_price if contract_price > 0 else 0
        return result



    current_month = date.today().replace(day=1)
    current_month_str = current_month.strftime("%Y-%m")
    weights = 3
    if end is not None and current_month_str > end:
        return {"message": "Ошибка в дате фильтра.\n Дата окончания в фильтре меньше текущей даты"}, 400

    sess = Session()
    try:
        account: PaymentAccount = (
            sess.query(PaymentAccount)
            .filter(PaymentAccount.is_active)
            .order_by(PaymentAccount.start_date.desc())
            .limit(1)
            .one()
        )
        account_balance = account.Money
        account_date = account.start_date
    except Exception as e:
        return {'message': 'ошибка'}, 500
    finally:
        sess.close()
    
    try:
        np_start_date = np.datetime64(account_date.strftime("%Y-%m"))

        if not end:
            np_end_date = np.datetime64(current_month_str) + np.timedelta64('13', 'M')
        else:
            np_end_date = np.datetime64(end) + np.timedelta64('1', 'M')
        end_date: date = (np_end_date + np.timedelta64(f'{weights-1}', 'M')).astype(date)
        date_to_remove: List[date] = np.arange(np_end_date, np_end_date + np.timedelta64(f'{weights+3}', 'M'), np.timedelta64('1', 'M')).astype(str)
        month_range = np.arange(
            np_start_date, 
            np_end_date + np.timedelta64(f'{weights-1}', 'M'), # дата окончания + длина тренда - 1 месяц начала - 1 последний месяц для тренда
            np.timedelta64('1', 'M')
        )
        months_all = {_.strftime("%Y-%m"): 0 for _ in month_range.astype(date)}
    except:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400

    months_plan = {_: 0 for _ in months_all if _ >= current_month_str}

    sess = Session()
    try:
        if current_month == account_date:
            current_balance = account_balance
        else:
            debit_fact: int = get_fact('Доход')
            credit_fact: int = get_fact('Расход')
            current_balance = account_balance + debit_fact - credit_fact
        
        debit_plan: Dict[str, int] = get_plan('Доход')
        credit_plan: Dict[str, int] = get_plan('Расход')
    except Exception as e:
        raise e
    finally:
        sess.close()
    
    

    dashboard = dict()
    for key, debit, credit in zip(debit_plan, debit_plan.values(), credit_plan.values()):
        current_balance = round(current_balance + debit - credit, 2)
        dashboard[key] = current_balance


    trend_value = {
        key: round(value, 2) 
        for key, value in zip(
            dashboard, 
            np.convolve(
                list(dashboard.values()), 
                np.ones(weights), 
                'valid'
            )/weights
        ) if key not in date_to_remove
    }
    
    for key in date_to_remove:
        if key in dashboard:
            dashboard.pop(key)

    result = {
        'months': dashboard,
        'trend': trend_value,
    }

    debit_on_current_date = (
            sess.query(func.sum(Invoice.payment_amount))
            .join(Contract)
            .filter(
                Contract.it_is == 'Доход',
                Contract.is_active,
                Invoice.payment_date < date.today(),
                Invoice.payment_date >= account_date,
            )
            .scalar()
        ) or 0

    credit_on_current_date = (
            sess.query(func.sum(Invoice.payment_amount))
            .join(Contract)
            .filter(
                Contract.it_is == 'Расход',
                Contract.is_active,
                Invoice.payment_date < date.today(),
                Invoice.payment_date >= account_date,
            )
            .scalar()
        ) or 0

    result['current_balance'] = account_balance + debit_on_current_date - credit_on_current_date
    
    result['info'] = {
        'current_balance': 'Текущий баланс.',
        'months': 'Прогноз остатка на первое число указанного месяца.<br>'
                  'Для расчёта остатка текущего (первого) месяца:<br>'
                  f'{TWO_SPACE}1. Берётся ближайшая прошедшая дата внесённого в систему остатка<br>'
                  f'{TWO_SPACE}2. Добавляются фактические доходы и расходы до первого числа текущего месяца<br>'
                  'Расчёт остатка оставшихся месяцев:<br>'
                  f'{TWO_SPACE}1. Берётся остаток предыдущего месяца<br>'
                  f'{TWO_SPACE}2. Добавляются плановые доходы и расходы до первого числа расчитываемого месяца<br>'
                  'Плановый доход/расход:<br>'
                  f'{TWO_SPACE}1. Ищет все контракты дохода/расход, у которых дата попадает в выбранный диапазон в следующих столбцах:<br>'
                  f'{FOUR_SPACE}а. Платёж получен: Дата<br>'
                  f'{FOUR_SPACE}б. Ожидаемая дата платежа<br>'
                  f'{FOUR_SPACE}в. Месяц платежа<br>'
                  f'{FOUR_SPACE}г. Дата окончания заказа<br>'
                  f'{TWO_SPACE}2. Если \"Месяц платежа\" или \"Дата окончания заказа\" входят в диапазон<br>'
                  f'{FOUR_SPACE}а. Общая стоимость заказа добавляется к общей сумме<br>'
                  f'{FOUR_SPACE}б. Счета ВНЕ выбранного диапазона вычитаются из общей суммы<br>'
                  f'{TWO_SPACE}3. Если \"Месяц платежа\" или \"Дата окончания заказа\" ВНЕ выбранного диапазона<br>'
                  f'{FOUR_SPACE}а. К общей сумме прибавляю только счета в выбранном диапазоне<br>'
                  'Правила вычитания счетов:<br>'
                  f'{TWO_SPACE}1. Если дата в столбце \"Платёж получен: Дата\" ВНЕ выбранного диапазона, то вычитается \"Платёж получен: Сумма\"<br>'
                  f'{TWO_SPACE}2. Если столбец \"Платёж получен: Дата\" пуст, а дата в столбце \"Ожидаемая дата платежа\" ВНЕ выбранного диапазона, то вычитается \"Счёт выставлен: Сумма\"<br>'
                  f'{TWO_SPACE}3. Все остальные счета игнорируются<br>'
                  'Правила прибавления счетов:<br>'
                  f'{TWO_SPACE}1. Если дата в столбце \"Платёж получен: Дата\" в выбранном диапазоне, то прибавляется \"Платёж получен: Сумма\"<br>'
                  f'{TWO_SPACE}2. Если столбец в столбце \"Платёж получен: Дата\" пуст, а дата в столбце \"Ожидаемая дата платежа\" в выбранном диапазоне, то прибавляется \"Счёт выставлен: Сумма\"<br>'
                  f'{TWO_SPACE}3. Все остальные счета игнорируются<br>'
    }

    return result, 200



def export_balance_forecast(end: str):
    """Получение контрактов за выбранный месяц, которые участвовали в расчёте остатка средств"""
    if end is None:
        return {'message': 'ошибка в датах. Нет даты окончания'}, 400
    if end.find('-') < 0:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400


    try:
        np_start_date = np.datetime64(end)
        np_end_date = np_start_date + np.timedelta64(1, 'M')
    except np.core._exceptions._UFuncInputCastingError:
        date_tuple = end.split('-')
        if len(date_tuple) == 3:
            np_start_date = np.datetime64(f'{date_tuple[0]}-{date_tuple[1]}')
            np_end_date = np_start_date + np.timedelta64(1, 'M')
        else:
            return {'message': 'ошибка в датах. Неодходим формат YYYY-MM'}, 400
    except ValueError:
        return {'message': 'ошибка в датах. Проверьте отправляемые даты'}, 400

    wb = Workbook()
    ws: Worksheet = wb.active

    ws.append((
        'ID',
        'тип',
        'месяц платежа',
        'контрагент',
        'Тип дохода/расхода',
        'Описание работ/продукции',
        'Номер договора',
        'Дата договора',
        'Номер заказа',
        'Дата заказа',
        'Дата окончания заказа',
        'Полученная сумма',
    )),
    # for cell in ws['C'][1:] + ws['H'][1:] + ws['J'][1:] + ws['K'][1:]:
    #     cell.number_format = 'dd.mm.yyyy'

    # for cell in ws['L'][1:]:
    #     cell.number_format = '##0.00 ₽'
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 27
    ws.column_dimensions['G'].width = 21
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 21
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 30


    start_date = np_start_date.astype(date)
    end_date = np_end_date.astype(date)

    sess = Session()
    try:
        plan_contract: List[Contract] = (
            sess.query(Contract)
            .join(Invoice, isouter=True)
            .filter(
                or_(
                    and_(
                        Contract.order_deadline >= start_date,
                        Contract.order_deadline < end_date,
                    ),
                    and_(
                        Contract.repeating_month >= start_date,
                        Contract.repeating_month < end_date,
                    ),
                    and_(
                        Invoice.payment_expected_date >= start_date,
                        Invoice.payment_expected_date < end_date,
                    ),
                    and_(
                        Invoice.payment_date >= start_date,
                        Invoice.payment_date < end_date,
                    ),
                ),
                Contract.is_active,
            )
        )
        for contract in plan_contract:
            if contract.repeating_month and start_date <= contract.repeating_month < end_date:
                base_price = True
                contract_price = contract.order_price
            elif not contract.repeating_month and contract.order_deadline and start_date <= contract.order_deadline < end_date:
                base_price = True
                contract_price = contract.order_price
            else:
                base_price = False
                contract_price = 0


            invoice: Invoice
            if base_price:
                for invoice in contract.fk__invoice:
                    contract_price -= invoice.price
                    if invoice.payment_date and start_date <= invoice.payment_date < end_date:
                        contract_price += invoice.payment_amount
                    elif not invoice.payment_date and start_date <= invoice.payment_expected_date < end_date:
                        contract_price += invoice.price
            else:
                for invoice in contract.fk__invoice:
                    if invoice.payment_date and start_date <= invoice.payment_date < end_date:
                        contract_price += invoice.payment_amount
                    elif not invoice.payment_date and start_date <= invoice.payment_expected_date < end_date:
                        contract_price += invoice.price
            
            if contract_price > 0:
                ws.append([
                    contract.id,
                    list(contract.it_is)[0],
                    contract.repeating_month,
                    contract.fk__contractor.name,
                    contract.fk__type.name,
                    contract.description,
                    contract.contract,
                    contract.contract_date,
                    contract.order,
                    contract.order_date,
                    contract.order_deadline,
                    contract_price,
                ])
        
    except Exception as e:
        raise e
    finally:
        sess.close()

    for cell in ws['C'][1:] + ws['H'][1:] + ws['J'][1:] + ws['K'][1:]:
        cell.number_format = 'dd.mm.yyyy'

    for cell in ws['L'][1:]:
        cell.number_format = '### ### ### ### ### ##0.00 ₽'

    bytes_book = BytesIO()
    wb.save(bytes_book)
    wb.close()
    bytes_book.seek(0)

    return bytes_book, 200
